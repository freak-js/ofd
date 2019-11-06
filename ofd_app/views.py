from django.contrib.auth import authenticate
from django.shortcuts import render
from django.http import QueryDict
from ofd_app.forms import ProductForm, UserForm, UserCreationFormCustom
from ofd_app.models import User, Product, ProductUserRel, Order, OrderStatus
from django.contrib.auth.models import Group
from django.contrib.auth.decorators import login_required, permission_required
from django import forms
from django.shortcuts import get_object_or_404
from django.shortcuts import redirect
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import login
from datetime import datetime, date, timedelta
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse, FileResponse
from ofd_app.filters import date_filter_format
from ofd_app.filters import apply_filters
from ofd_app.utils import *
from openpyxl import load_workbook
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from ofd_app.constants import *
from django.core.mail import send_mail
from ofd.settings import EMAIL_HOST_USER, TIME_ZONE
from django.utils.timezone import pytz
from django.template.loader import render_to_string
from weasyprint import HTML
from django.conf import settings
from ofd_app.number_to_text import num2text
import logging
from django.contrib.auth.signals import user_logged_in, user_logged_out, user_login_failed
from django.dispatch import receiver
from django.db.models.signals import pre_save


'''
Конфигурирование логгера из модуля logging
'''
logging.basicConfig(
    filename=PATH_TO_THE_LOGS, 
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    )


'''
@receiver(user_logged_in) - приемник сигнала и сигнал внутри.
user_logged_in - событие django об успешной авторизации пользователя.
signal_user_logged_in - логирует событие полученное приемником.
'''
@receiver(user_logged_in)
def signal_user_logged_in(sender, user, request, **kwargs):
    logging.info(f'id:{user.id}, username:{user.username} - успешно залогинился.')


'''
Аналагично предыдущему.
'''
@receiver(user_logged_out)
def signal_user_logged_out(sender, user, request, **kwargs):
    logging.info(f'id:{user.id}, username:{user.username} - разлогинился.')


'''
Аналагично предыдущему.
'''
@receiver(user_login_failed)
def signal_user_login_failed(sender, credentials, request, **kwargs):
    logging.warning(f'{credentials["username"]} - ошибка аутентификации!')


'''
signal_pre_save_change_password - отлавливает момент изменения пароля.
'''
@receiver(pre_save, sender=User)
def signal_pre_save_change_password(sender, **kwargs):
    user = kwargs.get('instance')

    if user:
        new_password = user.password
        try:
            old_password = User.objects.get(id=user.id).password
        except User.DoesNotExist:
            old_password = None

        if new_password != old_password:
            logging.info(f'id:{user.id}, username:{user.username} - изменил пароль.')


@login_required(login_url='/login/')
def product(request, **kwargs):

    if 'id' in kwargs:

        if not request.user.has_perm('ofd_app.change_product'):
            logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался изменить продукт id:{kwargs["id"]}!')
            return redirect('products')
        product = get_object_or_404(Product, product_id=kwargs['id'])

        if request.method == 'POST':
            form = ProductForm(request.POST, instance = product)

            if form.is_valid():
                form.save()
                logging.info(f'id:{request.user.id}, username:{request.user.username} - изменил продукт id:{kwargs["id"]}.')
        else:
            form = ProductForm(instance=product)
    else:

        if not request.user.has_perm('ofd_app.add_product'):
            logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался добавить новый продукт!')
            return redirect('products')

        if request.method == 'POST':
            form = ProductForm(request.POST)

            if form.is_valid():
                product = form.save()
                logging.info(f'id:{request.user.id}, username:{request.user.username} - добавил новый продукт id:{product.product_id}.')
        else:
            form = ProductForm()
    return render(request, 'ofd_app/index_product_add.html', {'form': form, 'user_role': request.user.get_role(), 'path': PRODUCTS})


@login_required(login_url='/login/')
@permission_required('ofd_app.view_product', login_url='/products/')
def products(request):

    if request.method == 'POST':
        product_id = to_int(request.POST.get('product_id', '').strip(), 0)
        order_comment = request.POST.get('order_comment', '').strip()
        quantity = to_int(request.POST.get('quantity', '').strip(), 1)

        if product_id > 0:
            product = request.user.get_product(product_id)
            cost = product.by_user__cost if product.by_user__cost is not None and product.by_user__cost > 0 else product.product_cost
            db_product = Product.objects.get(product_id=product_id)
            order = Order(user = request.user, product=db_product, comment = order_comment, amount = quantity, cost = cost, is_paid = False)
            order.save()
            logging.info(f'id:{request.user.id}, username:{request.user.username} - оформил заказ на продукт id:{product.product_id}.')
            send_mail(
                TEMPLATE_EMAIL_NEW_ORDER_ADMIN_SUBJECT, 
                TEMPLATE_EMAIL_NEW_ORDER_ADMIN_BODY.format(
                    first_name = request.user.first_name, 
                    last_name = request.user.last_name, 
                    corg=request.user.org, 
                    product = db_product.product_name, 
                    amount = quantity, 
                    total = quantity * cost 
                    ), 
                EMAIL_HOST_USER, 
                [EMAIL_HOST_USER], 
                fail_silently=False
                )
            logging.info(f'Отправлен email на {EMAIL_HOST_USER}.')
            send_mail(
                TEMPLATE_EMAIL_NEW_ORDER_USER_SUBJECT, 
                TEMPLATE_EMAIL_NEW_ORDER_USER_BODY.format(
                    number = order.id, 
                    date = order.adddate.astimezone(pytz.timezone(TIME_ZONE)).strftime("%Y.%m.%d %H:%M:%S"), 
                    product = db_product.product_name, 
                    amount = quantity, 
                    total = quantity * cost 
                    ), 
                EMAIL_HOST_USER, 
                [request.user.email], 
                fail_silently=False
                )
            logging.info(f'Отправлен email на {request.user.email}.')

        return redirect('products')

    return render(
        request, 
        'ofd_app/index_top.html', 
            {
            'products'  : request.user.get_products(), 
            'can_delete': request.user.has_perm('ofd_app.delete_product'), 
            'user_role' : request.user.get_role(), 
            'path'      : PRODUCTS
            })


@require_POST
@login_required(login_url='/login/')
@permission_required('ofd_app.delete_product', login_url='/products/')
def product_delete(request):
    ids = request.POST.getlist('product_to_delete')
    cnt_delete = 0
    for id in ids:
        try:
            product = Product.objects.get(product_id=id)
            product.product_is_active = False
            product.save()
            cnt_delete += 1
            logging.info(f'id:{request.user.id}, username:{request.user.username} - удалил продукт id:{id}.')
        except Product.DoesNotExist:
            logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался удалить несуществующий продукт!')
    return redirect('products')


@login_required(login_url='/login/')
def user(request, **kwargs):
    user = None
    my_card = False
    if 'id' in kwargs:
        my_card = request.user.id == kwargs['id']
        if not request.user.has_perm('ofd_app.change_user') and not my_card:
            return redirect('products')
        user = get_object_or_404(User, id=kwargs['id'])
        if not request.user.has_access_to_user(user):
            return redirect('products')
        if request.method == 'POST':
            user_form = UserForm(request.POST, instance = user, requested_user = request.user)
            user_save(user_form, request.user)
        else:
            user_form = UserForm(instance = user, requested_user = request.user)
    elif request.method == 'POST':
        if not request.user.has_perm('ofd_app.change_user'):
            return redirect('products')
        user_form = UserCreationFormCustom(request.POST, requested_user = request.user)
        user = user_save(user_form, request.user)
        if user is not None:
            return redirect('users')
    else:
        if not request.user.has_perm('ofd_app.view_user'):
            return redirect('products')
        user_form = UserCreationFormCustom(requested_user = request.user)
    return render(request, 'ofd_app/user.html', {'user_form': user_form, 'user_role': request.user.get_role(), 'path': MY_CARD if my_card else USERS})


@login_required(login_url='/login/')
@permission_required('ofd_app.change_productuserrel', login_url='/products/')
def user_product(request, **kwargs):
    if 'id' in kwargs:
        user = get_object_or_404(User, id=kwargs['id'])
        if not user.is_manager():
            logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался назначить цену не менеджеру!')
            return redirect('users')
        if request.method == 'POST':
            ProductUserRel.save_product_user_rel(request.POST, user, request.user.id)
        products = user.get_products()
        logging.info(f'id:{request.user.id}, username:{request.user.username} - назначил новую стоимость продуктам для id:{user.id}, username:{user.username}.')
    return render(request, 'ofd_app/user_product.html', {'products': products, 'path': USERS})


@login_required(login_url='/login/')
@permission_required('ofd_app.view_user', login_url='/products/')
def users(request):
    apply_filters(request, 'user_filters', {'org'})
    filters = {}
    if request.user.groups.filter(name='Manager').exists():
        users = User.objects.all().filter(is_active = True).filter(is_superuser=False).filter(parent=request.user)
    else:
        org = request.session['user_filters']['org']
        users = User.objects.all().filter(is_active = True).filter(is_superuser=False).filter(groups__name__in=['Manager', 'Admin'])
        if request.user.is_admin():
            users = users.filter(groups__name__in=['Manager'])
        if org is not None and len(org) > 0 and org != '*':
            users = users.filter(org=org)
        filters['org'] = User.get_organizations()
    user_data = []
    for user in users:
        data = user.__dict__
        data['role'] = user.get_role()
        data['childs'] = user.get_childs()
        user_data.append(data)
    return render(request, 'ofd_app/users.html', {'users': construct_pagination(request, user_data), 'can_delete': request.user.has_perm('ofd_app.delete_user'), 'filters': filters, 'user_role': request.user.get_role(), 'path': USERS})


@require_POST
@login_required(login_url='/login/')
@permission_required('ofd_app.delete_user', login_url='/products/')
def user_delete(request):
    ids = request.POST.getlist('user_to_delete')
    cnt_delete = 0
    for id in ids:
        if to_int(id, 0) > 0:
            try:
                user = User.objects.get(id=id)
                if request.user.has_access_to_user(user):
                    user.is_active = False
                    user.save()
                    cnt_delete += 1
                    logging.info(f'id:{request.user.id}, username:{request.user.username} - удалил пользователя id:{id}.')
            except User.DoesNotExist:
                logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался удалить несуществующего пользователя!')
    return redirect('users')


@login_required(login_url='/login/')
def orders(request):
    date = datetime.now()
    apply_filters(request, 'order_filters', {'date', 'status', 'org', 'user', 'paid'})
    if request.method == 'POST' and request.user.has_perm('ofd_app.manage_order_status'):

        id = to_int(request.POST.get('order_id', 0), 0)
        status = request.POST.get('status', '').strip()
        admin_comment = request.POST.get('admin_comment', '').strip()
        codes = request.POST.get('order_codes', '').strip()

        if id > 0 and len(status) > 0:
            try:
                order = Order.objects.get(id=id)
                op_result = order.assign_status(status, admin_comment, codes)

                if op_result:
                    logging.info(f'id:{request.user.id}, username:{request.user.username} - изменил статус заказа № МО-{order.id} на {status}.')
                    send_mail(
                        TEMPLATE_EMAIL_ORDER_STATUS_USER_SUBJECT, 
                        TEMPLATE_EMAIL_ORDER_STATUS_USER_BODY.format(
                            number = id, 
                            date = order.adddate.astimezone(pytz.timezone(TIME_ZONE)).strftime("%Y.%m.%d %H:%M:%S"), 
                            total = order.amount * order.cost, 
                            product = order.product.product_name, 
                            amount = order.amount, 
                            status = MESSAGES[0] if status == 'R' else MESSAGES[1], 
                            comment_template = 'Комментарий администратора: ' if admin_comment else '',
                            comment = admin_comment if admin_comment else ''
                            ), 
                        EMAIL_HOST_USER, 
                        [order.user.email], 
                        fail_silently=False
                        )
                    logging.info(f'Отправлен email на {order.user.email}.')
            except Order.DoesNotExist:
                logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался изменить статус несуществующего заказа!')

    date_from = datetime.strptime(request.session['order_filters']['date_from'], date_filter_format())
    date_to = datetime.strptime(request.session['order_filters']['date_to'], date_filter_format())
    org = request.session['order_filters']['org']
    status = request.session['order_filters']['status']
    user = request.session['order_filters']['user']
    paid = request.session['order_filters']['paid']
    orders = Order.get_orders(request.user, date_from, date_to, status, org, user, paid)
    order_data = []
    for order in orders:
        product = {
            'product_name': order.product.product_name, 
            'amount'      : order.amount, 'cost': order.cost, 
            'full_cost'   : order.amount * order.cost
            }
        order_data.append({
            'id'           : order.id, 
            'adddate'      : order.adddate.astimezone(pytz.timezone(TIME_ZONE)).strftime("%d.%m.%y %H:%M"), 
            'comment'      : order.comment, 
            'product'      : product, 
            'status'       : order.status.code, 
            'user'         : order.user, 
            'user_role'    : order.user_role, 
            'admin_comment': order.admin_comment, 
            'codes'        : order.codes, 
            'is_paid'      : order.is_paid,
            'allow_receipt': (DATE_AFTER_WHICH_INVOICES_WILL_BE_ISSUED - order.adddate).days < 0
        })
    filters = {}
    if request.user.is_superuser or request.user.is_admin():
        filters['org'] = User.get_organizations()
    elif request.user.is_manager():
        filters['users'] = [{'id': request.user.id, 'first_name': request.user.first_name, 'last_name': request.user.last_name}]
        for item in request.user.get_childs():
            child = {'id': item['id'], 'first_name': item['first_name'], 'last_name': item['last_name']}
            filters['users'].append(child)
    filters['status'] = OrderStatus.get_all_statuses()
    return render(request, 'ofd_app/orders.html', {'orders': construct_pagination(request, order_data), 'filters': filters, 'user_role': request.user.get_role(), 'path': ORDERS})


@login_required(login_url='/login/')
def stat_org(request):
    apply_filters(request, 'stat_org', {'date'})
    date_from = datetime.strptime(request.session['stat_org']['date_from'], date_filter_format())
    date_to = datetime.strptime(request.session['stat_org']['date_to'], date_filter_format()) + timedelta(1)
    sql = '''
    select 1 as id
         , u.org
         , u.inn
         , coalesce(sum(case when q.status = 'A' then q.total else 0 end), 0) as total
         , count(q.order_id) as cnt_all
         , sum(case when q.status = 'A' then 1 else 0 end) as cnt_approve
         , sum(case when q.status = 'I' then 1 else 0 end) as cnt_in_progress
         , sum(case when q.status = 'R' then 1 else 0 end) as cnt_reject
        from ofd_app_user u
            left outer join (
                select ug.user_id
                 from ofd_app_user_groups ug
                      inner join auth_group ag
                              on ug.group_id = ag.id
                where ag.name = 'Admin'
            ) ad on u.id = ad.user_id
            left outer join
            (
            select o.user_id
                 , o.id as order_id
                 , o.status_id as status
                 , o.amount * o.cost as total
              from ofd_app_order o
             where adddate >= %s
               and adddate < %s
            ) q on u.id = q.user_id
     where u.is_superuser = false
       and ad.user_id is null
    group by org, inn
    union
    select 1 as id
         , 'Общий итог' as org
         , '' as inn
         , coalesce(sum(case when o.status_id = 'A' then o.amount * o.cost else 0 end), 0) as total
         , count(o.id) as cnt_all
         , sum(case when o.status_id = 'A' then 1 else 0 end) as cnt_approve
         , sum(case when o.status_id = 'I' then 1 else 0 end) as cnt_in_progress
         , sum(case when o.status_id = 'R' then 1 else 0 end) as cnt_reject
      from ofd_app_order o
           inner join ofd_app_user u on o.user_id = u.id
           left outer join (
                select ug.user_id
                 from ofd_app_user_groups ug
                      inner join auth_group ag
                              on ug.group_id = ag.id
                where ag.name = 'Admin'
            ) ad on u.id = ad.user_id
     where u.is_superuser = false
       and ad.user_id is null
       and adddate >= %s
       and adddate < %s
    order by total desc
    '''
    result = User.objects.raw(sql, [date_from, date_to, date_from, date_to])
    data = []
    for row in result:
        item = {'org': row.org, 'inn': row.inn, 'total': row.total, 'cnt_all': row.cnt_all, 'cnt_approve': row.cnt_approve, 'cnt_in_progress': row.cnt_in_progress, 'cnt_reject': row.cnt_reject}
        data.append(item) 
    for i in data:
        if i['org'] == 'Общий итог':
            grand_total = i
    logging.info(f'id:{request.user.id}, username:{request.user.username} - смотрит статистику.')
    return render(request, 'ofd_app/stat_org.html', {'stat': construct_pagination(request, data), 'user_role': request.user.get_role(), 'path': STAT, 'grand_total' : grand_total})


@login_required(login_url='/login/')
def exportxlsx(request, **kwargs):
    if 'id' in kwargs:
        wb = Workbook()
        ws = wb.active
        db_codes = Order.get_order_codes(request.user, kwargs['id'])
        codes = db_codes.split()
        ix = 1
        for code in codes:
            col = 'A' + str(ix)
            ws[col] = code
            ix = ix + 1
        response = HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=codes.xlsx'
        logging.info(f'id:{request.user.id}, username:{request.user.username} - забрал XLSX с кодами к заказу № МО-{kwargs["id"]}.')
        return response


@login_required(login_url='/login/')
def exporttxt(request, **kwargs):
    if 'id' in kwargs:
        db_codes = Order.get_order_codes(request.user, kwargs['id'])
        codes = db_codes.split()
        response = HttpResponse('\r\n'.join(codes), content_type='text/plain')
        response['Content-Disposition'] = 'attachment; filename=codes.txt'
        logging.info(f'id:{request.user.id}, username:{request.user.username} - забрал TXT с кодами к заказу № МО-{kwargs["id"]}.')
        return response


def construct_pagination(request, data):
    page        = to_int(request.GET.get('page', 1), 1)
    p           = Paginator(data, 10)
    page_object = p.get_page(page)
    pagination  = {
                'page' : page if 0 < page <= p.num_pages else p.num_pages,
                'data' : page_object.object_list,
                'prev' : page_object.previous_page_number() if page_object.has_previous() else None,
                'next' : page_object.next_page_number() if page_object.has_next() else None,
                'count': get_pages_list(int(p.num_pages), page if 0 < page <= p.num_pages else p.num_pages)
                }
    return pagination


def user_reg(request):
    if request.method == 'POST':
        reg_form = UserCreationFormCustom(request.POST)
        user = user_save(reg_form)
        logging.info(f'id:{user.id}, username:{user.username} - успешно зарегестрировался.')
        if user is not None:
            send_mail(
                TEMPLATE_EMAIL_NEW_LOGIN_ADMIN_SUBJECT, 
                TEMPLATE_EMAIL_NEW_LOGIN_ADMIN_BODY.format(
                    first_name = user.first_name, 
                    last_name = user.last_name, 
                    email = user.email, 
                    phone_number = user.phone_number, 
                    org = user.org, 
                    city = user.city 
                    ), 
                EMAIL_HOST_USER, 
                [EMAIL_HOST_USER], 
                fail_silently=False
                )
            logging.info(f'Отправлен email на {EMAIL_HOST_USER}.')
            send_mail(
                TEMPLATE_EMAIL_NEW_LOGIN_USER_SUBJECT.format(first_name = user.first_name), 
                TEMPLATE_EMAIL_NEW_LOGIN_USER_BODY.format(login = user.username), 
                EMAIL_HOST_USER, 
                [user.email], 
                fail_silently=False
                )
            logging.info(f'Отправлен email на {user.email}.')
            login(request, user)
            return redirect('products')
    else:
        reg_form = UserCreationFormCustom()
    return render(request, 'ofd_app/user_reg.html', {'reg_form': reg_form})


def user_assign_group(user, group_name):
    group = Group.objects.get(name=group_name).user_set.add(user)


def user_resolve_group(user, request_user):
    group_name = 'Manager'

    if request_user is not None and request_user.groups.filter(name='Manager').exists():
        group_name = 'User'

    elif request_user is not None and request_user.is_superuser:
        group_name = 'Admin'
    user_assign_group(user, group_name)


def user_save(user_form, request_user=None):
    user = None

    if user_form.is_valid():
        user = user_form.save()

        if user.groups.all().count() == 0 and not user.is_superuser: #почему проверяется юзер из формы на суперюзера?
            user_resolve_group(user, request_user)

        if user.parent is None and user.is_user() and request_user is not None and request_user.is_manager():
            user.parent = request_user
            logging.info(f'id:{request_user.id}, username:{request_user.username} - добавил сотрудника {user.username}.')

        if user.is_user() and (user.inn is None or user.org is None or len(user.inn) == 0 or len(user.org) == 0):
            user.inn = user.parent.inn
            user.org = user.parent.org
        user.save()
    return user


'''
feedback - отдает страницу с контактами 
для прошедших аутентификацию пользователей.
'''
@login_required(login_url='/login/')
def feedback(request):
    logging.info(f'id:{request.user.id}, username:{request.user.username} - смотрит контакты.')
    return render(request, 'ofd_app/feedback.html', {'user_role': request.user.get_role(), 'path': FEEDBACK})


'''
instruction - отдает страницу с инструкцией.
'''
@login_required(login_url='/login/')
def instruction(request):
    logging.info(f'id:{request.user.username}, username:{request.user.username} - смотрит интсрукцию.')
    return render(request, 'ofd_app/instruction.html', {'user_role': request.user.get_role(), 'path': INSTRUCTION})


'''
contacts - отдает страницу с контактами со страницы авторизации
для незалогиненых, или не прошедших регистрацию пользователей.
'''
def contacts(request):
    logging.info(f'незалогиненый пользователь - смотрит контакты.')
    return render(request, 'ofd_app/contacts.html')


'''
Функция автоматического выставления счета
get_order_invoice - принимает request, возвращает http_response с готовым pdf
испльзует сторонние библиотеки: 
weasyprint(https://weasyprint.readthedocs.io/en/latest/install.html)
number_to_text(https://github.com/seriyps/ru_number_to_text)
зависимости и порядок импортов: 
    from django.template.loader import render_to_string
    from weasyprint import HTML
    from django.conf import settings
    from ofd_app.number_to_text import num2text.
'''
@login_required(login_url='/login/')
def get_order_invoice(request):
    order_id = to_int(request.POST.get('score_product_id', 0), 0)

    if order_id > 0:
        order = get_object_or_404(Order, id=order_id)

        if request.user.has_access_to_user(order.user):
            rendered_html = render_to_string('ofd_app/invoicing.html', context={
                'id'           : order.id, 
                'add_date'     : order.adddate.strftime("%Y.%m.%d"), 
                'amount'       : order.amount, 
                'org'          : order.user.org, 
                'inn'          : order.user.inn, 
                'product_name' : order.product.product_name, 
                'cost'         : '{0:,}'.format(order.cost).replace(',', ' ') + ',00', 
                'total'        : '{0:,}'.format(order.cost * order.amount).replace(',', ' ') + ',00', 
                'total_text'   : num2text(order.cost * order.amount),
                })
            pdf_file = HTML(string=rendered_html, base_url=request.build_absolute_uri()).write_pdf()
            http_response = HttpResponse(pdf_file, content_type='application/pdf')
            http_response['Content-Disposition'] = f'filename=MO-{order.id}'
            logging.info(f'id:{request.user.id}, username:{request.user.username} - забрал счет к заказу № МО-{order_id}.')
            return http_response
    logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался получить доступ к счету заказа № МО-{order_id}!')
    return redirect('orders')


'''
Функция автоматического выставления УПД
get_upd - полностью идентична по алгоритму работы и зависимостям 
с get_order_invoice кроме сторонней библиотеки number_to_text.
'''
@login_required(login_url='/login/')
def get_upd(request):
    order_id = to_int(request.POST.get('score_product_id', 0), 0)

    if order_id > 0:
        order = get_object_or_404(Order, id=order_id)

        if request.user.has_access_to_user(order.user):
            rendered_html = render_to_string('ofd_app/upd.html', context={
                'id'             : order.id, 
                'add_date'       : order.adddate.strftime("%d.%m.%Y"),
                'date_of_payment': order.date_of_payment.strftime("%d.%m.%Y"),
                'amount'         : order.amount, 
                'org'            : order.user.org, 
                'inn'            : order.user.inn, 
                'product_name'   : order.product.product_name, 
                'cost'           : '{0:,}'.format(order.cost).replace(',', ' ') + ',00', 
                'total'          : '{0:,}'.format(order.cost * order.amount).replace(',', ' ') + ',00', 
                })
            pdf_file = HTML(string=rendered_html, base_url=request.build_absolute_uri()).write_pdf()
            http_response = HttpResponse(pdf_file, content_type='application/pdf')
            http_response['Content-Disposition'] = f'filename=UPD-MO-{order.id}'
            logging.info(f'id:{request.user.id}, username:{request.user.username} - забрал УПД к заказу № МО-{order_id}.')
            return http_response
    logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался получить доступ к УПД заказа № МО-{order_id}!')
    return redirect('orders')


'''
order_change_pay_sign - функция изменения статуса оплаты заказа.
Доступна только админу и году.
Принимает список id заказов и меняет bool значение Order.is_paid на противоположное.
'''
@require_POST
@login_required(login_url='/login/')
def order_change_pay_sign(request):
    if not (request.user.is_superuser or request.user.is_admin()):
        logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался изменить статус оплаты!')
        return redirect('orders')

    ids = request.POST.getlist('is_paid')

    for id in ids:
        if to_int(id, 0) > 0:
            try:
                order = Order.objects.get(id=id)
                order.is_paid = not order.is_paid
                order.save()
                logging.info(f'id:{request.user.id}, username:{request.user.username} - изменение статуса оплаты заказа № МО-{id} на {order.is_paid}.')
            except Order.DoesNotExist:
                logging.warning(f'id:{request.user.id}, username:{request.user.username} - попытался изменить статус оплаты несуществующего заказа!')
    return redirect('orders')


'''
Функция изменения стоимости и/или количества индивидуального заказа.
change_order - принимает из request словарь с ключами: cost, amount, order_id.
Производит обновление соответсвующих ключам (cost, amount) полей в БД.
'''
@require_POST
@login_required(login_url='/login/')
def change_order(request):
    if not (request.user.is_superuser or request.user.is_admin()):
        logging.warning(f'id:{request.user.id}, username:{request.user.username} - попытался воспользоваться change_order!')
        return redirect('orders')

    new_cost   = to_int(request.POST.get('cost', '').strip(), 0)
    new_amount = to_int(request.POST.get('amount', '').strip(), 0)
    order_id   = to_int(request.POST.get('order_id', '').strip(), 0)

    if order_id > 0 and new_amount > 0 and new_cost > 0:
        try:
            order = Order.objects.get(id=order_id)
        except Order.DoesNotExist:
            return redirect('orders')
        order.amount = new_amount
        order.cost   = new_cost
        order.save(update_fields=['amount', 'cost'])
        logging.info(f'id:{request.user.id}, username:{request.user.username} - изменил заказ № МО-{order.id}.')
    return redirect('orders')


'''
logs - функция отдающая фронтенду 4 отсортированных 
списка с разным уровнем логирования событий при помощи sort_logs().
Каждый список включает в себя 100 последних записей для быстрого анализа,
или мониторинга состояния прилодения.
'''
@login_required(login_url='/login/')
def logs(request):

    if not (request.user.is_superuser or request.user.is_admin()):
        logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался просмотреть логи!')
        return redirect('orders')

    lists = sort_logs()
    logging.info(f'id:{request.user.id}, username:{request.user.username} - смотрит логи.')
    return render(
        request, 
        'ofd_app/logs.html', {
            'info'          : lists[0][:100], 
            'warning'       : lists[1][:100], 
            'error'         : lists[2][:100], 
            'critical'      : lists[3][:100],
            'info_count'    : len(lists[0]), 
            'warning_count' : len(lists[1]), 
            'error_count'   : len(lists[2]), 
            'critical_count': len(lists[3])
            })


'''
get_log_file - отдает админу, или суперпользователю файл формата .txt
'''
@login_required(login_url='/login/')
def get_log_file(request):

    if not (request.user.is_superuser or request.user.is_admin()):
        logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался скачать логи!')
        return redirect('orders')
    
    logging.info(f'id:{request.user.id}, username:{request.user.username} - скачал логи.')
    return FileResponse(open(PATH_TO_THE_LOGS, 'rb'),  as_attachment = True, filename=f'logs_{datetime.today().strftime("%m_%d_%Y")}.log')


'''
add_codes - принимает из request файл формата xlsx и id продута,
читает данные из ячеек, вносит изменения в соответсвующие таблицы базы.
Встретив пустую ячейку - прекращает считывание и завершает цикл.
Выполнив более 500 итераций прекращает считывание, логирует ошибку 
и высылает уведомление на email.
'''
@require_POST
@login_required(login_url='/login/')
def add_codes(request):
    
    if not (request.user.is_superuser or request.user.is_admin()):
        logging.warning(f'id:{request.user.id}, username:{request.user.username} - пытался загрузить на сервер коды!')
        return redirect('orders')

    try:
        xlsx_file = request.FILES.get('new_codes_file')
        wb = openpyxl.load_workbook(xlsx_file)
        ws = wb.active
    except:
        logging.warning(f'id:{request.user.id}, username:{request.user.username} - ошибка с загружаемым файлом!')
        return render(request, 'ofd_app/results.html', {'exception': True})

    product_id = to_int(request.POST.get('product_id'))

    if product_id:
        iteration = 2 #Начинаем с двойки из-за названий колонок в первых ячейках

        while iteration <= 501:
            cell_A = ws[f'A{iteration}'].value
            cell_B = ws[f'B{iteration}'].value

            if cell_A is None or cell_B is None:
                break

            # Выполняем проверку на количество итераций на случай не явного наличия данных в ячейках документа.

            if iteration > 500:
                logging.critical('Ошибка в процессе загрузки кодов! Загружено более 500 кодов, или не были соблюдены условия выхода из цикла!')
                send_mail('Ошибка при загрузке кодов!', 'Загружено более 500 кодов, или не были соблюдены условия выхода из цикла!', EMAIL_HOST_USER, [EMAIL_HOST_USER], fail_silently=False)
                return render(request, 'ofd_app/results.html', {'exception': True})

            iteration += 1

    return render(request, 'ofd_app/results.html', {'exception': False, 'iteration': iteration - 1})
